import csv
from pathlib import Path
from typing import T

from openpyxl import load_workbook


def csv_to_dict(file: Path) -> list[dict[str, str]]:
    res = []
    reding = open(file)
    data = csv.reader(reding)
    head = next(data)
    for line in data:
        res.append(
            {head[i]: line[i] for i in range(len(head))}
        )
    return res


def dict_to_csv(data: list[dict[str, str]], file: Path):
    writer = csv.DictWriter(open(file, "w", newline=""), fieldnames=data[0].keys())
    writer.writeheader()
    for row in data:
        writer.writerow(row)


def write_file(content: str, path: Path):
    path.write_text(content)


def read_file(path: Path) -> str:
    return path.read_text()


def xlsx_to_dict(file_path: str | Path) -> list[dict[str, str]]:
    workbook = load_workbook(file_path)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if headers[i] is not None}
        data.append(row_dict)

    return data

